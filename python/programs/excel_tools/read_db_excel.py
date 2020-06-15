#读取数据库D模型信息
import openpyxl
import logging

from table_info import TableInfo
from field_info import FieldInfo
from parm_info import ParmInfo

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

def read_db_table_info(db_table_excel):
    """从数据库表中读取表名、字段等信息
    参数：
        db_table_excel: 数据库D模型excel路径
    返回值：
        TableInfo: 据D模型excel读取的数据库表和字段信息对象
    """
    wb = openpyxl.load_workbook(db_table_excel)
    sheet = wb['表结构']
    filed_list = []
    for row in range(2, sheet.max_row + 1):
        field_eng_nm = sheet['I' + str(row)].value
        if None == field_eng_nm:
            continue
        field_chn_nm = sheet['J' + str(row)].value
        filed_dict_no = sheet['K' + str(row)].value
        filed_db_type = sheet['L' + str(row)].value
        filed_length = sheet['M' + str(row)].value
        
        if 'P' == sheet['O' + str(row)].value:
            filed_is_key = True
        else:
            filed_is_key = False
        
        if '是' == sheet['R' + str(row)].value:
            filed_is_nullable = True
        else:
            filed_is_nullable = False
        
        field_info = FieldInfo(field_eng_nm, field_chn_nm, filed_dict_no, 
        filed_db_type, filed_length, filed_is_key, filed_is_nullable,
        '', '', '', '')
        filed_list.append(field_info)
    
    table_eng_nm = sheet['G2'].value
    table_chn_nm = sheet['H2'].value
    table_info = TableInfo(table_eng_nm, table_chn_nm, filed_list)
    logging.info('数据库表名：' + table_info.table_eng_nm + " " + table_info.table_chn_nm)
    logging.info('字段个数：' + str(len(table_info.fields_info)))
<<<<<<< Updated upstream
    return table_info
=======
    return table_info

def get_parm_infos(field_infos):
    table_info_dict = {}
    parm_table_dict = {}
    for field_info in field_infos:
        if field_info.table_eng_nm in table_info_dict:
            table_info_dict[field_info.table_eng_nm].fields_info.append(field_info)
        else:
            fields = [field_info]
            table_info = TableInfo(field_info.table_eng_nm, field_info.table_chn_nm, fields)
            table_info_dict[field_info.table_eng_nm] = table_info
        
        if field_info.parm_no in parm_table_dict:
            parm_table_dict[field_info.parm_no].add(field_info.table_eng_nm)
        else:
            table_set = set()
            table_set.add(field_info.table_eng_nm)
            parm_table_dict[field_info.parm_no] = table_set
    
    parm_infos = []
    for parm_no, table_set in parm_table_dict.items():
        table_infos = []
        for table_eng_nm in table_set:
            table_infos.append(table_info_dict[table_eng_nm])
        field_info = table_infos[0].fields_info[0]
        parm_info = ParmInfo(parm_no, field_info.parm_nm, table_infos)
        parm_infos.append(parm_info)
    return parm_infos



def read_export_db_excel(export_db_excel):
    """从数据库SQL导出的表中读取表名、字段等信息
    参数：
        export_db_excel: 从数据库SQL导出的表
    返回值：
        TableInfo:  从数据库SQL导出的表读取的数据库表和字段信息对象
    """
    wb = openpyxl.load_workbook(export_db_excel)
    sheet = wb.worksheets[0]
    filed_list = []
    for row in range(2, sheet.max_row + 1):
        field_eng_nm = sheet['C' + str(row)].value
        if None == field_eng_nm:
            continue
        field_chn_nm = sheet['D' + str(row)].value
        filed_dict_no = sheet['E' + str(row)].value
        filed_db_type = sheet['F' + str(row)].value
        filed_length = sheet['G' + str(row)].value
        
        if 'P' == sheet['H' + str(row)].value:
            filed_is_key = True
        else:
            filed_is_key = False
        
        if '是' == sheet['I' + str(row)].value:
            filed_is_nullable = True
        else:
            filed_is_nullable = False
        
        table_eng_nm = sheet['A' + str(row)].value
        table_chn_nm = sheet['B' + str(row)].value
        parm_no = sheet['J' + str(row)].value
        parm_nm = sheet['K' + str(row)].value
        
        field_info = FieldInfo(field_eng_nm, field_chn_nm, filed_dict_no, 
        filed_db_type, filed_length, filed_is_key, filed_is_nullable, 
        table_eng_nm, table_chn_nm, parm_no, parm_nm)
        filed_list.append(field_info)
    
    # table_eng_nm = sheet['A2'].value
    # table_chn_nm = sheet['B2'].value
    # table_info = TableInfo(table_eng_nm, table_chn_nm, filed_list)
    # logging.info('数据库表名：' + table_info.table_eng_nm + " " + table_info.table_chn_nm)
    logging.info('字段个数：' + str(len(filed_list)))
    parm_infos = get_parm_infos(filed_list)
    logging.info('参数个数：' + str(len(parm_infos)))

    return parm_infos

def read_table_info(db_excel, mode):
    if 'export' == mode:
        return read_export_db_excel(db_excel)
    else:
        return read_db_table_info(db_excel)
>>>>>>> Stashed changes
